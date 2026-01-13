"""
Script para importar datos del Excel original a la aplicación
"""
import openpyxl
import json
import os
from pathlib import Path
from datetime import datetime
import re

EXCEL_FILE = '2025-2026 Nablijven.xlsx'
OUTPUT_DIR = Path('data')

# Crear directorio de datos si no existe
OUTPUT_DIR.mkdir(exist_ok=True)

def import_students():
    """Importa estudiantes de las hojas 'Leerlingen'"""
    wb = openpyxl.load_workbook(EXCEL_FILE)
    students = []
    
    day_mapping = {
        'Leerlingen MAANDAG': 'MAANDAG',
        'Leerlingen DINSDAG': 'DINSDAG',
        'Leerlingen DONDERDAG': 'DONDERDAG',
    }
    
    for sheet_name, day in day_mapping.items():
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in range(1, ws.max_row + 1):
                name = ws.cell(row=row, column=1).value
                grade = ws.cell(row=row, column=2).value
                
                if name and name.strip():
                    student = {
                        'id': f'student-{day}-{row}',
                        'name': str(name).strip(),
                        'grade': str(grade).strip() if grade else '',
                        'day': day
                    }
                    students.append(student)
    
    output_file = OUTPUT_DIR / 'students.json'
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(students, f, ensure_ascii=False, indent=2)
    
    print(f"[OK] Importados {len(students)} estudiantes a {output_file}")
    wb.close()

def import_detentions():
    """Importa detenciones de todas las hojas de sesiones"""
    wb = openpyxl.load_workbook(EXCEL_FILE)
    detentions = []
    
    # Mapeo de días de la semana
    day_mapping = {
        'MA': 'MAANDAG',
        'Di': 'DINSDAG',
        'DI': 'DINSDAG',
        'Do': 'DONDERDAG',
        'DO': 'DONDERDAG',
    }
    
    for sheet_name in wb.sheetnames:
        # Ignorar hojas de estudiantes y plantillas
        if sheet_name.startswith('Leerlingen') or sheet_name.startswith('Sjabloon'):
            continue
        
        # Determinar día de la semana
        day_of_week = None
        for prefix, day in day_mapping.items():
            if sheet_name.startswith(prefix):
                day_of_week = day
                break
        
        if not day_of_week:
            continue
        
        ws = wb[sheet_name]
        
        # Intentar parsear la fecha del nombre de la hoja
        # Formato esperado: "MA 8 sept", "DI 9 sept", "DO 11 sept", etc.
        session_date = None
        date_patterns = [
            (r'(\d{1,2})\s+(sept|okt|nov|dec|jan|feb|maart|apr|mei|jun|jul|aug)', re.IGNORECASE),
        ]
        
        month_map = {
            'sept': 9, 'okt': 10, 'nov': 11, 'dec': 12,
            'jan': 1, 'feb': 2, 'maart': 3, 'apr': 4,
            'mei': 5, 'jun': 6, 'jul': 7, 'aug': 8
        }
        
        for pattern, flags in date_patterns:
            match = re.search(pattern, sheet_name, flags)
            if match:
                day = int(match.group(1))
                month_str = match.group(2).lower()
                month = month_map.get(month_str, 9)  # Default a septiembre
                # Asumir año 2025 (ajustar si es necesario)
                year = 2025 if month >= 9 else 2026
                try:
                    session_date = datetime(year, month, day).strftime('%Y-%m-%d')
                    break
                except ValueError:
                    pass
        
        # Si no se pudo parsear, usar fecha por defecto
        if not session_date:
            session_date = '2025-09-01'
        
        for row in range(3, ws.max_row + 1):
            number = ws.cell(row=row, column=1).value
            student = ws.cell(row=row, column=2).value
            teacher = ws.cell(row=row, column=3).value
            reason = ws.cell(row=row, column=4).value
            task = ws.cell(row=row, column=5).value
            lvs_date = ws.cell(row=row, column=6).value
            should_print = ws.cell(row=row, column=7).value
            can_chromebook = ws.cell(row=row, column=8).value
            extra_notes = ws.cell(row=row, column=9).value
            
            # Solo procesar si hay un estudiante
            if student and str(student).strip():
                # Convertir fecha LVS si es datetime
                lvs_date_str = ''
                if lvs_date:
                    if isinstance(lvs_date, datetime):
                        lvs_date_str = lvs_date.strftime('%Y-%m-%d')
                    else:
                        lvs_date_str = str(lvs_date)
                
                detention = {
                    'id': f'detention-{sheet_name}-{row}',
                    'number': int(number) if number else row - 2,
                    'date': session_date,
                    'dayOfWeek': day_of_week,
                    'student': str(student).strip(),
                    'teacher': str(teacher).strip() if teacher else '',
                    'reason': str(reason).strip() if reason else '',
                    'task': str(task).strip() if task else '',
                    'lvsDate': lvs_date_str,
                    'shouldPrint': bool(should_print) if should_print is not None else False,
                    'canUseChromebook': bool(can_chromebook) if can_chromebook is not None else False,
                    'extraNotes': str(extra_notes).strip() if extra_notes else '',
                }
                detentions.append(detention)
    
    output_file = OUTPUT_DIR / 'detentions.json'
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(detentions, f, ensure_ascii=False, indent=2)
    
    print(f"[OK] Importadas {len(detentions)} detenciones a {output_file}")
    print("[INFO] Nota: Revisa las fechas importadas, especialmente las que no se pudieron parsear del nombre de la hoja")
    wb.close()

if __name__ == '__main__':
    print("Importando datos del Excel...")
    print("-" * 50)
    import_students()
    import_detentions()
    print("-" * 50)
    print("[OK] Importacion completada")
